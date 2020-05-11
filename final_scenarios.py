# -*- coding: utf-8 -*-
"""
Created on Wed Apr 22 12:05:37 2020
@author: bernardino, francesco, maria
"""
from openpyxl import load_workbook
import pandas as pd
# function to convert the xlsx sheet into a 2D array
def iter_rows(ws):
    Arr = []
    for row in ws.iter_rows():
        result = []
        for cell in row:
            result.append(cell.value)
        Arr.append(result)
    return Arr

def findSheet(ID, Arr):
    for i in range(1, len(Arr)):
        if ID == Arr[i][0]:
            return Arr[i][3]
        
def find_CShare_And_PopShare(ID, Arr):
    for i in range(1, len(Arr)):
        if ID == Arr[i][0]:
            # return column C "Share GDP" and column G "Share (Population) decimal"
            return float(Arr[i][2]), float(Arr[i][5])
        
        
# ---READ emissions_reduction_master_sheet.xlsx FILE----------------------
Loc = r'/Users/Maria/Desktop/Process_Google_Stats_COVID-19/emissions_reduction_master_sheet.xlsx'
wb = load_workbook(filename=Loc)
sheet_RegShare = wb['Regional_shares']
ArrayData_sheet_RegShare = iter_rows(sheet_RegShare)
sheet_modal_trans_countries_covered = wb['modal_trans_countries_covered']
ArrayData_sheet_modal_trans = iter_rows(sheet_modal_trans_countries_covered)
##---READ Global_Mobility_Report.csv FILE----------------------
Loc = '/Users/Maria/Desktop/Process_Google_Stats_COVID-19/Global_Mobility_Report.csv'
Array = pd.read_csv(Loc, sep=',', header=None)
Array_Mobility_Data = Array.fillna(0.)
Array_Mobility_Data = Array_Mobility_Data.values
##----------------------------------------------------------------
# ------------------------ TO BE UPDATED AS NEW REPORTS GET RElEASED --------- #
#Date = "2020-04-26"
Date = "26/04/2020"
ID_country_list = ["AW",
                   "AF",
                   "AO",
                   "AI",
                   "AL",
                   "AD",
                   "AG",
                   "AE",
                   "AR",
                   "AM",
                   "AU",
                   "AT",
                   "AZ",
                   "BI",
                   "BE",
                   "BJ",
                   "BF",
                   "BD",
                   "BG",
                   "BH",
                   "BS",
                   "BA",
                   "BY",
                   "BZ",
                   "BM",
                   "BO",
                   "BR",
                   "BB",
                   "BN",
                   "BT",
                   "BW",
                   "CF",
                   "CA",
                   "CH",
                   "CL",
                   "CN",
                   "CI",
                   "CM",
                   "CD",
                   "CG",
                   "CK",
                   "CO",
                   "KM",
                   "CV",
                   "CR",
                   "ZZZZ",
                   "CU",
                   "CW",
                   "KY",
                   "CY",
                   "CZ",
                   "DE",
                   "DJ",
                   "DM",
                   "DK",
                   "DO",
                   "ZZZZ",
                   "ZZZZ",
                   "DZ",
                   "EC",
                   "EG",
                   "ER",
                   "ES",
                   "EE",
                   "ET",
                   "FI",
                   "FJ",
                   "FR",
                   "FM",
                   "GA",
                   "GB",
                   "GE",
                   "GH",
                   "GN",
                   "GM",
                   "GW",
                   "GQ",
                   "GR",
                   "GD",
                   "GL",
                   "GT",
                   "GY",
                   "HK",
                   "HN",
                   "HR",
                   "HT",
                   "HU",
                   "ID",
                   "IN",
                   "IE",
                   "IR",
                   "IQ",
                   "IS",
                   "IL",
                   "IT",
                   "JM",
                   "JO",
                   "JP",
                   "KZ",
                   "KE",
                   "KG",
                   "KH",
                   "KI",
                   "KN",
                   "KR",
                   "ZZZZ",
                   "KW",
                   "LA",
                   "LB",
                   "LR",
                   "LY",
                   "LC",
                   "LI",
                   "LK",
                   "LS",
                   "LT",
                   "LU",
                   "LV",
                   "MO",
                   "MA",
                   "MC",
                   "MD",
                   "MG",
                   "MV",
                   "MX",
                   "MH",
                   "MK",
                   "ML",
                   "MT",
                   "MM",
                   "ME",
                   "MN",
                   "MZ",
                   "MR",
                   "MS",
                   "MU",
                   "MW",
                   "MY",
                   "NA",
                   "NC",
                   "NE",
                   "NG",
                   "NI",
                   "NL",
                   "NO",
                   "NP",
                   "NR",
                   "NZ",
                   "OM",
                   "PK",
                   "PA",
                   "PE",
                   "PH",
                   "PW",
                   "PG",
                   "PL",
                   "PR",
                   "KP",
                   "PT",
                   "PY",
                   "PS",
                   "PF",
                   "QA",
                   "RO",
                   "RU",
                   "RW",
                   "SA",
                   "SD",
                   "SS",
                   "SN",
                   "SG",
                   "SB",
                   "SL",
                   "SV",
                   "SM",
                   "SO",
                   "ST",
                   "SD",
                   "SR",
                   "SK",
                   "SI",
                   "SE",
                   "SZ",
                   "SX",
                   "SC",
                   "SY",
                   "TC",
                   "TD",
                   "TG",
                   "TH",
                   "TJ",
                   "TM",
                   "TL",
                   "TO",
                   "TT",
                   "TN",
                   "TR",
                   "TV",
                   "TW",
                   "TZ",
                   "ZZZZ",
                   "UG",
                   "UA",
                   "UY",
                   "US",
                   "ZZZZ",
                   "UZ",
                   "VC",
                   "VE",
                   "VG",
                   "VN",
                   "VU",
                   "WS",
                   "YE",
                   "ZZZZ",
                   "ZZZZ",
                   "ZA",
                   "ZM",
                   "ZW"]

CountriesCovered_list = []
Retail_Recreation_list = []
Grocery_Farmacy_list = []
Parks_list = []
Transit_Stat_list = []
Workplaces_list = []
Residential_list = []
car_modal_share = []
car_vehicle_share = []

for country in ID_country_list:
    for row in range(1, len(Array_Mobility_Data)):
        if str(Array_Mobility_Data[row][0]) == country:
            if str(Array_Mobility_Data[row][4]) == Date:
                if int(Array_Mobility_Data[row][2]) == 0:
                    CountriesCovered_list.append(str(country))
                    Retail_Recreation_list.append(float(Array_Mobility_Data[row][5]))
                    Grocery_Farmacy_list.append(float(Array_Mobility_Data[row][6]))
                    Parks_list.append(float(Array_Mobility_Data[row][7]))
                    Transit_Stat_list.append(float(Array_Mobility_Data[row][8]))
                    Workplaces_list.append(float(Array_Mobility_Data[row][9]))
                    Residential_list.append(float(Array_Mobility_Data[row][10]))
                    print(Array_Mobility_Data[row][1], float(Array_Mobility_Data[row][5]), float(Array_Mobility_Data[row][6]),float(Array_Mobility_Data[row][7]), float(Array_Mobility_Data[row][8]),float(Array_Mobility_Data[row][9]),float(Array_Mobility_Data[row][10]))
                    break
                
# ---------Initialise regions' sheets in emissions_reduction_master_sheet.xlsc with zeros--------#
def initialiseColumns():
    for country in CountriesCovered_list:
        region = findSheet(country, ArrayData_sheet_RegShare)
        if region != str(0) or region != str(0.):
#            val = CountriesCovered_list.index(country)
#            print(val, str(region), country)
            sheet = wb[str(region)]
            for i in range(2, 28):
                sheet["B" + str(i)] = 0.
initialiseColumns()

# ----------Read values for modal transport share and car share over total vehicles-------------#
for country in CountriesCovered_list:
    for row in ArrayData_sheet_modal_trans:
        if row[0] == country:
            print(country, float(row[3]), row[7])
            car_modal_share.append(float(row[3]))
            #car_vehicle_share.append(float(row[7]))
            car_vehicle_share.append(float(row[7]))
            

# -------------------------------------------Useful functions-------------------------------------#
def saveExcel(name):
    wb.save(name + '.xlsx')   
    
def columnB(region, weight, val, scenario):
    sheet = wb[str(region)]
    if scenario == 1:
        sheet["B21"] = float(iter_rows(sheet)[20][1]) + (float(Retail_Recreation_list[val] + 0. * Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B20"] = float(iter_rows(sheet)[19][1]) + (float(Retail_Recreation_list[val]) / 100.) * weight
        sheet["B27"] = float(iter_rows(sheet)[26][1]) + (float(Retail_Recreation_list[val]) / 100.) * weight
        sheet["B6"] = float(iter_rows(sheet)[5][1]) + (float(0. * Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B10"] = float(iter_rows(sheet)[9][1]) + (float(0. * Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B25"] = float(iter_rows(sheet)[24][1]) + (float(Parks_list[val]) / 100.) * weight
        sheet["B19"] = float(iter_rows(sheet)[18][1]) + (float(Transit_Stat_list[val]) / 100.) * weight
        sheet["B16"] = float(iter_rows(sheet)[15][1]) + (float(Workplaces_list[val]) / 100.) * weight
        sheet["B17"] = float(iter_rows(sheet)[16][1]) + (float(Workplaces_list[val]) / 100.) * weight
        sheet["B16"] = float(iter_rows(sheet)[15][1]) + (float(Residential_list[val]) / 100.) * weight
        sheet["B17"] = float(iter_rows(sheet)[16][1]) + (float(Residential_list[val]) / 100.) * weight
        max_loss = min(float(Retail_Recreation_list[val]), float(Workplaces_list[val]))
        sheet["B8"] = float(iter_rows(sheet)[7][1]) + (max_loss / 100.) * weight * car_modal_share[val] * car_vehicle_share[val]
    elif scenario == 2:
        sheet["B21"] = float(iter_rows(sheet)[20][1]) + (float(Retail_Recreation_list[val] + 0.25 * Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B20"] = float(iter_rows(sheet)[19][1]) + (float(Retail_Recreation_list[val]) / 100.) * weight
        sheet["B27"] = float(iter_rows(sheet)[26][1]) + (float(Retail_Recreation_list[val]) / 100.) * weight
        sheet["B6"] = float(iter_rows(sheet)[5][1]) + (float(0.25 * Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B10"] = float(iter_rows(sheet)[9][1]) + (float(0.25 * Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B25"] = float(iter_rows(sheet)[24][1]) + (float(Parks_list[val]) / 100.) * weight
        sheet["B19"] = float(iter_rows(sheet)[18][1]) + (float(Transit_Stat_list[val]) / 100.) * weight
        sheet["B16"] = float(iter_rows(sheet)[15][1]) + (float(Workplaces_list[val]) / 100.) * weight
        sheet["B17"] = float(iter_rows(sheet)[16][1]) + (float(Workplaces_list[val]) / 100.) * weight
        sheet["B16"] = float(iter_rows(sheet)[15][1]) + (float(Residential_list[val]) / 100.) * weight
        sheet["B17"] = float(iter_rows(sheet)[16][1]) + (float(Residential_list[val]) / 100.) * weight
        max_loss = min(float(Retail_Recreation_list[val]), float(Workplaces_list[val]))
        sheet["B8"] = float(iter_rows(sheet)[7][1]) + (max_loss / 100.) * weight * car_modal_share[val] * car_vehicle_share[val]
    elif scenario == 3:
        sheet["B21"] = float(iter_rows(sheet)[20][1]) + (float(Retail_Recreation_list[val] + 0.5 * Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B20"] = float(iter_rows(sheet)[19][1]) + (float(Retail_Recreation_list[val]) / 100.) * weight
        sheet["B27"] = float(iter_rows(sheet)[26][1]) + (float(Retail_Recreation_list[val]) / 100.) * weight
        sheet["B6"] = float(iter_rows(sheet)[5][1]) + (float(0.5 * Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B10"] = float(iter_rows(sheet)[9][1]) + (float(0.5 * Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B25"] = float(iter_rows(sheet)[24][1]) + (float(Parks_list[val]) / 100.) * weight
        sheet["B19"] = float(iter_rows(sheet)[18][1]) + (float(Transit_Stat_list[val]) / 100.) * weight
        sheet["B16"] = float(iter_rows(sheet)[15][1]) + (float(Workplaces_list[val]) / 100.) * weight
        sheet["B17"] = float(iter_rows(sheet)[16][1]) + (float(Workplaces_list[val]) / 100.) * weight
        sheet["B16"] = float(iter_rows(sheet)[15][1]) + (float(Residential_list[val]) / 100.) * weight
        sheet["B17"] = float(iter_rows(sheet)[16][1]) + (float(Residential_list[val]) / 100.) * weight
        max_loss = min(float(Retail_Recreation_list[val]), float(Workplaces_list[val]))
        sheet["B8"] = float(iter_rows(sheet)[7][1]) + (max_loss / 100.) * weight * car_modal_share[val] * car_vehicle_share[val]
    elif scenario == 4:
        sheet["B21"] = float(iter_rows(sheet)[20][1]) + (float(Retail_Recreation_list[val] + 0.75 * Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B20"] = float(iter_rows(sheet)[19][1]) + (float(Retail_Recreation_list[val]) / 100.) * weight
        sheet["B27"] = float(iter_rows(sheet)[26][1]) + (float(Retail_Recreation_list[val]) / 100.) * weight
        sheet["B6"] = float(iter_rows(sheet)[5][1]) + (float(0.75 * Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B10"] = float(iter_rows(sheet)[9][1]) + (float(0.75 * Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B25"] = float(iter_rows(sheet)[24][1]) + (float(Parks_list[val]) / 100.) * weight
        sheet["B19"] = float(iter_rows(sheet)[18][1]) + (float(Transit_Stat_list[val]) / 100.) * weight
        sheet["B16"] = float(iter_rows(sheet)[15][1]) + (float(Workplaces_list[val]) / 100.) * weight
        sheet["B17"] = float(iter_rows(sheet)[16][1]) + (float(Workplaces_list[val]) / 100.) * weight
        sheet["B16"] = float(iter_rows(sheet)[15][1]) + (float(Residential_list[val]) / 100.) * weight
        sheet["B17"] = float(iter_rows(sheet)[16][1]) + (float(Residential_list[val]) / 100.) * weight
        max_loss = min(float(Retail_Recreation_list[val]), float(Workplaces_list[val]))
        sheet["B8"] = float(iter_rows(sheet)[7][1]) + (max_loss / 100.) * weight * car_modal_share[val] * car_vehicle_share[val]
    elif scenario == 5:
        sheet["B21"] = float(iter_rows(sheet)[20][1]) + (float(Retail_Recreation_list[val] + Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B20"] = float(iter_rows(sheet)[19][1]) + (float(Retail_Recreation_list[val]) / 100.) * weight
        sheet["B27"] = float(iter_rows(sheet)[26][1]) + (float(Retail_Recreation_list[val]) / 100.) * weight
        sheet["B6"] = float(iter_rows(sheet)[5][1]) + (float(Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B10"] = float(iter_rows(sheet)[9][1]) + (float(Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B25"] = float(iter_rows(sheet)[24][1]) + (float(Parks_list[val]) / 100.) * weight
        sheet["B19"] = float(iter_rows(sheet)[18][1]) + (float(Transit_Stat_list[val]) / 100.) * weight
        sheet["B16"] = float(iter_rows(sheet)[15][1]) + (float(Workplaces_list[val]) / 100.) * weight
        sheet["B17"] = float(iter_rows(sheet)[16][1]) + (float(Workplaces_list[val]) / 100.) * weight
        sheet["B16"] = float(iter_rows(sheet)[15][1]) + (float(Residential_list[val]) / 100.) * weight
        sheet["B17"] = float(iter_rows(sheet)[16][1]) + (float(Residential_list[val]) / 100.) * weight
        max_loss = min(float(Retail_Recreation_list[val]), float(Workplaces_list[val]))
        sheet["B8"] = float(iter_rows(sheet)[7][1]) + (max_loss / 100.) * weight * car_modal_share[val] * car_vehicle_share[val]
    elif scenario == 6:
        sheet["B21"] = float(iter_rows(sheet)[20][1]) + (float(Retail_Recreation_list[val] + Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B20"] = float(iter_rows(sheet)[19][1]) + (float(Retail_Recreation_list[val]) / 100.) * weight
        sheet["B27"] = float(iter_rows(sheet)[26][1]) + (float(Retail_Recreation_list[val]) / 100.) * weight
        sheet["B6"] = float(iter_rows(sheet)[5][1]) + (float(Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B10"] = float(iter_rows(sheet)[9][1]) + (float(Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B25"] = float(iter_rows(sheet)[24][1]) + (float(0. * Parks_list[val]) / 100.) * weight
        sheet["B19"] = float(iter_rows(sheet)[18][1]) + (float(Transit_Stat_list[val]) / 100.) * weight
        sheet["B16"] = float(iter_rows(sheet)[15][1]) + (float(Workplaces_list[val]) / 100.) * weight
        sheet["B17"] = float(iter_rows(sheet)[16][1]) + (float(Workplaces_list[val]) / 100.) * weight
        sheet["B16"] = float(iter_rows(sheet)[15][1]) + (float(Residential_list[val]) / 100.) * weight
        sheet["B17"] = float(iter_rows(sheet)[16][1]) + (float(Residential_list[val]) / 100.) * weight
        max_loss = min(float(Retail_Recreation_list[val]), float(Workplaces_list[val]))
        sheet["B8"] = float(iter_rows(sheet)[7][1]) + (max_loss / 100.) * weight * car_modal_share[val] * car_vehicle_share[val]
    elif scenario == 7:
        sheet["B21"] = float(iter_rows(sheet)[20][1]) + (float(Retail_Recreation_list[val] + 0. * Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B20"] = float(iter_rows(sheet)[19][1]) + (float(Retail_Recreation_list[val]) / 100.) * weight
        sheet["B27"] = float(iter_rows(sheet)[26][1]) + (float(Retail_Recreation_list[val]) / 100.) * weight
        sheet["B6"] = float(iter_rows(sheet)[5][1]) + (float(0. * Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B10"] = float(iter_rows(sheet)[9][1]) + (float(0. * Grocery_Farmacy_list[val]) / 100.) * weight
        sheet["B25"] = float(iter_rows(sheet)[24][1]) + (float(0. * Parks_list[val]) / 100.) * weight
        sheet["B19"] = float(iter_rows(sheet)[18][1]) + (float(Transit_Stat_list[val]) / 100.) * weight
        sheet["B16"] = float(iter_rows(sheet)[15][1]) + (float(Workplaces_list[val]) / 100.) * weight
        sheet["B17"] = float(iter_rows(sheet)[16][1]) + (float(Workplaces_list[val]) / 100.) * weight
        sheet["B16"] = float(iter_rows(sheet)[15][1]) + (float(Residential_list[val]) / 100.) * weight
        sheet["B17"] = float(iter_rows(sheet)[16][1]) + (float(Residential_list[val]) / 100.) * weight
        max_loss = min(float(Retail_Recreation_list[val]), float(Workplaces_list[val]))
        sheet["B8"] = float(iter_rows(sheet)[7][1]) + (max_loss / 100.) * weight * car_modal_share[val] * car_vehicle_share[val]
    else:
        print("WRONG SCENARIO")
    
def runForEachCountry(flag, name, scenario):
    initialiseColumns()
    for country in CountriesCovered_list:
        region = findSheet(country, ArrayData_sheet_RegShare)
        if region != str(0) or region != str(0.):
            country_weight, popShare = find_CShare_And_PopShare(country, ArrayData_sheet_RegShare)
            val = CountriesCovered_list.index(country)
            if flag == 1: columnB(region, country_weight, val, scenario)
            else: columnB(region, popShare, val, scenario)
    saveExcel(name)

# ----------------------------------------Run for all scenarios-------------------------------------#
# Flag (1 - country_weight, 2 - popShare),  File name,              Scenario
#runForEachCountry(1,                        "S1_GDP",                   1)
runForEachCountry(2,                        "S1_POP",                   1)
#runForEachCountry(1,                        "S2_GDP",                   2)
runForEachCountry(2,                        "S2_POP",                   2)
#runForEachCountry(1,                        "S3_GDP",                   3)
#runForEachCountry(2,                        "S3_POP",                   3)
#runForEachCountry(1,                        "S4_GDP",                   4)
#runForEachCountry(2,                        "S4_POP",                   4)
#runForEachCountry(1,                        "S5_GDP",                   5)
#runForEachCountry(2,                        "S5_POP",                   5)
#runForEachCountry(1,                        "S6_GDP",                   6)
#runForEachCountry(2,                        "S6_POP",                   6)
runForEachCountry(1,                        "S7_GDP",                   7)
runForEachCountry(2,                        "S7_POP",                   7)